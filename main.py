import io
import json
import time

import openpyxl
import ctypes
from selenium import webdriver
import register


def load_excel(excel_path):
    excel_content = list()
    try:
        with open(excel_path, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
    except IOError:
        return False

    wb = openpyxl.load_workbook(in_mem_file, read_only=True)
    st = wb.worksheets[0]
    max_row = st.max_row
    max_col = st.max_column
    for i in range(max_row - 1):
        excel_content.append([])
        for j in range(max_col):
            excel_content[i].append(st.cell(row=i + 2, column=j + 1).value)
            j = j + 1
        i = i + 1

    return excel_content


def initialize_operation(speed):
    # browser_url = ""
    options = webdriver.ChromeOptions()
    # options.binary_location = browser_url
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(chrome_options=options)

    # Refresh rate
    driver.implicitly_wait(speed)

    driver.get("https://m.qlchat.com/video/admin/live/home")
    with open("cookies.json", "r") as f:
        cookies = json.load(f)
        for cookie in cookies:
            driver.add_cookie(cookie)
    driver.get("https://m.qlchat.com/video/admin/live/home")

    # Save cookies to json file
    if len(driver.find_elements_by_class_name("page-login")):
        ctypes.windll.user32.MessageBoxW(0, "Cookies已过期，请扫码重新登陆", "提示", 0x1000)
        while len(driver.find_elements_by_class_name("page-login")):
            time.sleep(5)

        cookies = driver.get_cookies()
        with open("cookies.json", "w") as f:
            json.dump(cookies, f)

    return driver


def create_course(driver, create_info):
    driver.get(create_info[1])

    element = driver.find_elements_by_class_name("edition-update-footer")
    if len(element):
        element[0].find_element_by_class_name("ant-btn").click()

    driver.find_elements_by_class_name("item")[2].click()
    driver.find_element_by_css_selector("[class='ant-btn ant-btn-primary ant-btn-lg']").click()
    # 课程主题
    driver.find_element_by_class_name("ant-input").send_keys(create_info[2])
    driver.find_element_by_class_name("btn-panel").find_element_by_css_selector("[class='ant-btn ant-btn-primary "
                                                                                "ant-btn-lg']").click()
    # 去优化
    driver.find_element_by_class_name("primary-btn").click()
    element_upload = driver.find_elements_by_class_name("file-input-container")
    # 上传课程图片
    if create_info[3] is not None:
        element_upload[0].find_element_by_tag_name("input").send_keys(create_info[3])
        time.sleep(2)
    # 主讲人
    if create_info[4] is not None:
        element_send = driver.find_elements_by_class_name("ant-form-item-children")
        element_send[4].find_element_by_tag_name("input").send_keys(create_info[4])
    # 主讲人介绍
    if create_info[5] is not None:
        element_send[5].find_element_by_tag_name("input").send_keys(create_info[5])

    # 上传直播概要图片
    if create_info[6] is not None:
        driver.switch_to.frame(driver.find_elements_by_tag_name("iframe")[0])    # 切换到新的frame
        driver.find_element_by_xpath("//input[@accept='image/*']").send_keys(create_info[6])
        time.sleep(2)
        driver.switch_to.default_content()    # 切换到默认主文档
        driver.find_element_by_xpath("//*[@class='btn confirm on-log']").click()
    
    # 上传视频
    element_upload[1].find_element_by_tag_name("input").send_keys(create_info[7])
    # 等待上传完毕
    while "100" not in driver.find_element_by_class_name("status").text:
        time.sleep(5)
        print(driver.find_element_by_class_name("status").text)
    # 保存
    driver.find_element_by_css_selector("[class='ant-btn ant-btn-primary ant-btn-lg']").click()
    # 结束完一个新建课程任务，等待时间
    time.sleep(5)


def start():
    driver = initialize_operation(5)
    qianliao_config = load_excel("qianliao_config.xlsm")
    for one_line in qianliao_config:
        print("正在上传：" + one_line[2])
        create_course(driver, one_line)
        print("上传完毕")


if __name__ == "__main__":
    reg = register.Register()
    if reg.check_authored():
        start()
